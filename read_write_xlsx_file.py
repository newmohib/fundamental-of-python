# Pypi and Pip (Huh)
# Python Packages
# install package go pypi.org and search package and install like this "pip install openpyxl"

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    # wb = xl.load_workbook("transactions.xlsx")
    sheet = wb['Sheet1']

    # total Row in Sheet1
    print(sheet.max_row)

    # last row is exclude in thar reason we added +1
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price_cell = sheet.cell(row, 4)

        if row == 1:
            corrected_price_cell.value = 'Corrected_Price'

        else:
            if not cell.value:
                break

            print('else', cell.value)
            corrected_price = float(cell.value) * 0.9
            corrected_price_cell.value = corrected_price

    # wb.save('transactions2.xlsx')

    # create a chart
    # do not create actual chart
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook("transactions.xlsx")
