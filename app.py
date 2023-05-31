
# Pypi and Pip (Huh)
# Python Packages
# install package go pypi.org and search package and install like this "pip install openpyxl"

import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['a1']
cell = sheet.cell(1, 1)

# print(cell.value)

# total Row in Sheet1
# print(sheet.max_row)

# last row is exclude in thar reason we added +1
for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

    # cell = sheet.cell(row, row)
    # value = cell.value
    # print(row)
    # print(cell)
    # print(value)
wb.save('transactions2.xlsx')
